import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO

# Caminho do arquivo Excel
arquivo_excel = 'seuarquivo.xlsx'  # altere para o nome correto
pasta_saida = 'imagens_exportadas'
coluna_base = 'A'  # Coluna que define o nome da imagem

# Cria a pasta de saída
os.makedirs(pasta_saida, exist_ok=True)

# Carrega a planilha
wb = load_workbook(arquivo_excel)
ws = wb.active

# Processa as imagens
for idx, img in enumerate(ws._images):
    # Identifica a célula de ancoragem
    if hasattr(img.anchor, '_from'):
        pos = img.anchor._from
        linha = pos.row + 1  # Corrige para 1-based index do Excel
        celula_nome = f"{coluna_base}{linha}"
        nome = ws[celula_nome].value or f"imagem_{idx+1}"
    else:
        nome = f"imagem_{idx+1}"

    # Define nome do arquivo com extensão .jpg
    nome_arquivo = f"{nome}.jpg"

    # Extrai e salva a imagem como JPG de alta qualidade
    if hasattr(img, '_data'):
        image_data = img._data()
        pil_image = PILImage.open(BytesIO(image_data))
        
        # Converter para RGB caso tenha canal alfa
        if pil_image.mode in ('RGBA', 'P'):
            pil_image = pil_image.convert('RGB')
        
        pil_image.save(os.path.join(pasta_saida, nome_arquivo), format='JPEG', quality=95)

print("Imagens exportadas como JPG em alta qualidade!")
